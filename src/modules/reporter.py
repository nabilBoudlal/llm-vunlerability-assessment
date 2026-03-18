"""
Risk Reporter — generates .xlsx vulnerability assessment reports.
"""
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK_NAVY = "1F2937"
C_RED       = "DC2626";  C_LIGHT_RED = "FEE2E2"
C_ORANGE    = "EA580C";  C_LIGHT_ORG = "FED7AA"
C_YELLOW    = "D97706";  C_LIGHT_YEL = "FEF3C7"
C_GREEN     = "16A34A";  C_LIGHT_GRN = "DCFCE7"
C_BLUE      = "2563EB";  C_LIGHT_BLU = "DBEAFE"
C_WHITE     = "FFFFFF"
C_LINK      = "1D4ED8"   # blue for hyperlinks

NVD_BASE_URL = "https://nvd.nist.gov/vuln/detail/"

SEV_CONFIG = {
    "Critical":      (C_RED,    C_LIGHT_RED, "🔴"),
    "High":          (C_ORANGE, C_LIGHT_ORG, "🟠"),
    "Medium":        (C_YELLOW, C_LIGHT_YEL, "🟡"),
    "Low":           (C_GREEN,  C_LIGHT_GRN, "🟢"),
    "Informational": (C_BLUE,   C_LIGHT_BLU, "ℹ️"),
}

def _border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col, value, bg=C_DARK_NAVY, fg=C_WHITE, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=fg, size=size)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c

def _cell(ws, row, col, value, bg=C_WHITE, bold=False, color="111827",
          align="left", wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=color, size=9)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _border()
    return c

def _cve_cell(ws, row, col, cve_string: str, bg=C_WHITE):
    """
    Render CVE IDs with NVD hyperlinks.
    - Single CVE  → clickable hyperlink to NVD detail page
    - Multiple    → first CVE is the hyperlink, rest appended as plain text
    - 'None' / '' → plain 'N/A' cell
    """
    if not cve_string or cve_string.strip().lower() in ("none", "n/a", ""):
        return _cell(ws, row, col, "N/A", bg=bg, align="left")

    # Split and clean CVE IDs
    cve_ids = [c.strip() for c in re.split(r"[,;\s]+", cve_string)
               if re.match(r"CVE-\d{4}-\d+", c.strip())]

    if not cve_ids:
        return _cell(ws, row, col, cve_string, bg=bg, align="left")

    primary = cve_ids[0]
    rest    = cve_ids[1:]

    display_text = primary if not rest else f"{primary} (+{len(rest)} more)"
    full_text    = ", ".join(cve_ids)          # tooltip / wrap fallback

    c = ws.cell(row=row, column=col, value=display_text)
    c.hyperlink  = f"{NVD_BASE_URL}{primary}"
    c.font       = Font(name="Arial", size=9, color=C_LINK,
                        underline="single", bold=False)
    c.fill       = PatternFill("solid", start_color=bg)
    c.alignment  = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border     = _border()
    # Store full CVE list in the cell comment-style via the value if multiple
    if rest:
        # Write full list as cell value so it's visible when cell is expanded
        c.value = full_text
        c.font  = Font(name="Arial", size=9, color=C_LINK,
                       underline="single", bold=False)
    return c


# ── Convert structured findings (ReAct output) to the data dict ───────────────
def _from_findings(findings: list, target: str) -> dict:
    """
    Convert the structured findings list produced by VAAgentReAct.run()
    into the same 'data' dict format that the sheet builders expect.

    findings item shape (from va_agent_react.py):
      {
        "port", "service", "target", "severity", "cvss",
        "cves",           # comma-separated string OR list
        "cve_refs",       # list of {id, url, cvss, has_exploit, actively_exploited}
        "analysis", "remediation",
        "has_exploit", "actively_exploited"
      }
    """
    data = {"target": target, "services": [], "policies": [], "remediations": []}

    for f in findings:
        port    = str(f.get("port", "?"))
        service = f.get("service", "unknown")
        sev     = f.get("severity", "Informational")
        cvss    = str(f.get("cvss", "N/A"))
        cves_raw = f.get("cves", "")
        # cves may arrive as a list or comma-separated string
        if isinstance(cves_raw, list):
            cves_str = ", ".join(cves_raw) if cves_raw else "None"
        else:
            cves_str = cves_raw or "None"

        exploit = "Yes" if f.get("has_exploit") or f.get("actively_exploited") else "No"

        data["services"].append({
            "port":     port,
            "service":  service,
            "severity": sev,
            "cvss":     cvss,
            "cves":     cves_str,
            "cve_refs": f.get("cve_refs", []),
            "policy":   "No",
            "exploit":  exploit,
            "analysis": f.get("analysis", ""),
        })

        # Build remediation entry
        rem_text = f.get("remediation", "")
        if rem_text:
            # "step1 | step2 | step3" → list of steps
            steps = [s.strip() for s in rem_text.split("|") if s.strip()]
            if steps:
                data["remediations"].append({"port": port, "steps": steps})

    # Sort services: Critical first
    order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Informational": 4}
    data["services"].sort(key=lambda x: order.get(x["severity"], 5))

    return data


def _parse(report_text: str, target: str) -> dict:
    data = {"target": target, "services": [], "policies": [], "remediations": []}

    inv = re.search(r"## 2\. Service Inventory.*?\n(.*?)(?=\n## |\Z)", report_text, re.S)
    if inv:
        for line in inv.group(1).splitlines():
            cols = [c.strip() for c in line.strip("|").split("|")]
            if len(cols) >= 7 and re.match(r"\d+", cols[0]):
                port, service, sev, cvss, cves, policy, exploit = cols[:7]
                sev_clean = "Informational"
                for s in ["Critical", "High", "Medium", "Low", "Informational"]:
                    if s.lower() in sev.lower():
                        sev_clean = s
                        break
                data["services"].append({
                    "port":    port.strip(),
                    "service": service.strip(),
                    "severity":sev_clean,
                    "cvss":    cvss.strip(),
                    "cves":    cves.strip(),
                    "policy":  "Yes" if "⚠️" in policy else "No",
                    "exploit": "Yes" if "Yes" in exploit or "💥" in exploit else "No",
                })

    analysis_map = {}
    for block in re.finditer(
        r"### [🔴🟠🟡🟢ℹ️]+ Port (\d+)[^\n]*\n.*?\*\*Analysis:\*\*\s*(.+?)(?=\n---|\n###|\Z)",
        report_text, re.S
    ):
        analysis_map[block.group(1)] = block.group(2).strip()

    for svc in data["services"]:
        svc["analysis"] = analysis_map.get(svc["port"], "")

    pol = re.search(r"## 3\. Policy.*?\n(.*?)(?=\n## |\Z)", report_text, re.S)
    if pol:
        for block in re.finditer(
            r"\*\*[^|]*Port (\d+)[^*]*\*\*[^\n]*\n+>\s*(.+?)(?=\n\*\*|\Z)",
            pol.group(1), re.S
        ):
            data["policies"].append({
                "port": block.group(1),
                "description": block.group(2).strip().replace("\n", " ")
            })

    rem = re.search(r"## 4\. Remediation Plan\n(.*?)(?=\n## |\Z)", report_text, re.S)
    if rem:
        for block in re.finditer(
            r"\*\*Port (\d+)[^*]+\*\*[^\n]*\n((?:\*[^\n]+\n?)+)",
            rem.group(1)
        ):
            steps = re.findall(r"\*\s+(.+)", block.group(2))
            data["remediations"].append({
                "port":  block.group(1),
                "steps": steps
            })

    return data


# ── Sheet builders ────────────────────────────────────────────────────────────
def _dashboard(wb, data, generated):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Vulnerability Assessment Report — {data['target']}"
    c.font = Font(name="Arial", bold=True, size=16, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Generated: {generated}   |   Framework: LLM-VA ReAct Agent"
    c.font = Font(name="Arial", size=9, color="6B7280")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    counts = {s: 0 for s in SEV_CONFIG}
    policy_count = exploit_count = 0
    for svc in data["services"]:
        if svc["severity"] in counts: counts[svc["severity"]] += 1
        if svc["policy"]  == "Yes": policy_count  += 1
        if svc["exploit"] == "Yes": exploit_count += 1

    ws.row_dimensions[3].height = 10
    cards = [
        ("Services Analysed", str(len(data["services"])), C_DARK_NAVY),
        ("Critical",          str(counts["Critical"]),    C_RED),
        ("High",              str(counts["High"]),         C_ORANGE),
        ("Medium",            str(counts["Medium"]),       C_YELLOW),
        ("Low",               str(counts["Low"]),          C_GREEN),
        ("Policy Alerts",     str(policy_count),           "7C3AED"),
        ("Public Exploits",   str(exploit_count),          "DC2626"),
    ]
    for i, (label, val, bg) in enumerate(cards, 1):
        for r in (4, 5, 6): ws.row_dimensions[r].height = [16, 32, 8][r-4]
        lc = ws.cell(row=4, column=i, value=label)
        lc.font = Font(name="Arial", size=8, bold=True, color=C_WHITE)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=5, column=i, value=val)
        vc.font = Font(name="Arial", size=22, bold=True, color=C_WHITE)
        vc.fill = PatternFill("solid", start_color=bg)
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[7].height = 10
    for i, h in enumerate(["Severity", "Count", "% of Total"], 1):
        _hdr(ws, 8, i, h)
    ws.row_dimensions[8].height = 20
    total = len(data["services"]) or 1
    for r, (sev, (color, bg, _)) in enumerate(SEV_CONFIG.items(), 9):
        _cell(ws, r, 1, sev, bg=bg, bold=True, color=color, align="center")
        _cell(ws, r, 2, counts[sev], align="center")
        pct = ws.cell(row=r, column=3)
        pct.value = f"=B{r}/{total}"
        pct.number_format = "0.0%"
        pct.font = Font(name="Arial", size=9)
        pct.alignment = Alignment(horizontal="center")
        pct.border = _border()
        ws.row_dimensions[r].height = 18

    for i, w in enumerate([22, 10, 12, 12, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _inventory(wb, data):
    ws = wb.create_sheet("Service Inventory")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Service Inventory — All Open Ports"
    c.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = ["Port", "Service / Version", "Severity", "CVSS", "CVE(s)",
            "Policy Alert", "Public Exploit", "LLM Analysis"]
    for i, h in enumerate(hdrs, 1):
        _hdr(ws, 2, i, h)
    ws.row_dimensions[2].height = 22

    for r, svc in enumerate(data["services"], 3):
        sev = svc["severity"]
        color, bg, emoji = SEV_CONFIG.get(sev, (C_BLUE, C_WHITE, ""))
        _cell(ws, r, 1, svc["port"], align="center")
        _cell(ws, r, 2, svc["service"])
        _cell(ws, r, 3, f"{emoji} {sev}", bg=bg, bold=True, color=color, align="center")
        _cell(ws, r, 4, svc["cvss"], align="center")
        _cve_cell(ws, r, 5, svc["cves"])                    # ← hyperlinked CVEs
        pol_bg = C_LIGHT_RED if svc["policy"] == "Yes" else C_WHITE
        _cell(ws, r, 6, svc["policy"], bg=pol_bg, align="center",
              color=C_RED if svc["policy"] == "Yes" else "111827")
        exp_bg = C_LIGHT_RED if svc["exploit"] == "Yes" else C_WHITE
        _cell(ws, r, 7, svc["exploit"], bg=exp_bg, align="center")
        _cell(ws, r, 8, svc.get("analysis", ""), wrap=True)
        ws.row_dimensions[r].height = 20

    for i, w in enumerate([7, 28, 16, 9, 40, 13, 14, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _findings(wb, data):
    ws = wb.create_sheet("Critical & High Findings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Critical & High Risk Findings"
    c.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = ["Port", "Service", "Severity", "CVSS", "CVE(s)", "Policy", "Analysis"]
    for i, h in enumerate(hdrs, 1):
        _hdr(ws, 2, i, h)
    ws.row_dimensions[2].height = 22

    r = 3
    for svc in data["services"]:
        if svc["severity"] not in ("Critical", "High"):
            continue
        sev = svc["severity"]
        color, bg, emoji = SEV_CONFIG.get(sev, (C_BLUE, C_WHITE, ""))
        _cell(ws, r, 1, svc["port"], align="center")
        _cell(ws, r, 2, svc["service"])
        _cell(ws, r, 3, f"{emoji} {sev}", bg=bg, bold=True, color=color, align="center")
        _cell(ws, r, 4, svc["cvss"], align="center")
        _cve_cell(ws, r, 5, svc["cves"])                    # ← hyperlinked CVEs
        pol_bg = C_LIGHT_RED if svc["policy"] == "Yes" else C_WHITE
        _cell(ws, r, 6, svc["policy"], bg=pol_bg, align="center",
              color=C_RED if svc["policy"] == "Yes" else "111827")
        _cell(ws, r, 7, svc.get("analysis", ""), wrap=True)
        ws.row_dimensions[r].height = 22
        r += 1

    for i, w in enumerate([7, 28, 16, 9, 40, 12, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _policies(wb, data):
    ws = wb.create_sheet("Policy Violations")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "Policy & Configuration Violations"
    c.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill("solid", start_color="7C3AED")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(["Port", "Service", "Violation Description"], 1):
        _hdr(ws, 2, i, h)
    ws.row_dimensions[2].height = 22

    port_map = {s["port"]: s["service"] for s in data["services"]}
    for r, pol in enumerate(data["policies"], 3):
        _cell(ws, r, 1, pol["port"], align="center")
        _cell(ws, r, 2, port_map.get(pol["port"], ""))
        _cell(ws, r, 3, pol["description"], bg=C_LIGHT_YEL, wrap=True)
        ws.row_dimensions[r].height = 30

    for i, w in enumerate([7, 28, 85], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _remediation(wb, data):
    ws = wb.create_sheet("Remediation Plan")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Remediation Plan"
    c.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    c.fill = PatternFill("solid", start_color=C_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(["Port", "Service", "Severity", "Recommended Actions"], 1):
        _hdr(ws, 2, i, h)
    ws.row_dimensions[2].height = 22

    port_map = {s["port"]: s["service"]  for s in data["services"]}
    sev_map  = {s["port"]: s["severity"] for s in data["services"]}

    for r, rem in enumerate(data["remediations"], 3):
        port = rem["port"]
        sev  = sev_map.get(port, "Informational")
        color, bg, emoji = SEV_CONFIG.get(sev, (C_BLUE, C_WHITE, ""))
        _cell(ws, r, 1, port, align="center")
        _cell(ws, r, 2, port_map.get(port, ""))
        _cell(ws, r, 3, f"{emoji} {sev}", bg=bg, bold=True, color=color, align="center")
        steps_text = "\n".join(f"• {s}" for s in rem["steps"])
        _cell(ws, r, 4, steps_text, bg=C_LIGHT_GRN, wrap=True)
        ws.row_dimensions[r].height = max(35, 18 * len(rem["steps"]))

    for i, w in enumerate([7, 28, 16, 90], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Public API ────────────────────────────────────────────────────────────────
class RiskReporter:
    def __init__(self, output_dir="reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def save_report(self, report_name: str,
                    report_content,          # list[dict] (ReAct) OR str (pipeline)
                    cve_sources: dict = None) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"{report_name}_report_{timestamp}.xlsx"
        filepath  = os.path.join(self.output_dir, filename)

        target = (report_name
                  .replace("_react_assessment", "")
                  .replace("_final_assessment", "")
                  .replace("_", "."))

        # ── Detect input format ───────────────────────────────────────────────
        if isinstance(report_content, list):
            # ReAct agent: structured findings list
            data = _from_findings(report_content, target)
        else:
            # Pipeline v6: markdown string
            data = _parse(report_content, target)
        generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = Workbook()
        _dashboard(wb, data, generated)
        _inventory(wb, data)
        _findings(wb, data)
        _policies(wb, data)
        _remediation(wb, data)

        wb.save(filepath)
        return filepath